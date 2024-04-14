# importanto da bibliotecas que serão usadas
import matplotlib.pyplot as plt
import openpyxl as xlsx
import pandas as pd
import json
import csv
'''
- csv-->json-->xls-->análise das planilhas e elaboração de relatórios
- primeiro, seleciona o arquivo .csv baixado das respostas do google
- segundo, traduz esse arquivo em formato json e altera os nomes para melhorar a análise
- terceiro, elabora gráficos, faz cálculos e adiquire todos os dados para a análise
- quarto, transforma tudo em uma planilha do excel que conterá todo o resultado
'''
#alterando nome das perguntas para facilitar a construção das colunas no excel
df = pd.read_json('teste3.json')
novas_keys={
    'Resposta 1':"Banana 1", 
    "Resposta 2":'Banana 2',
    "Resposta 3":"Banana 3"
}
df.rename(columns=novas_keys, inplace=True)
df.to_json('teste3.json', orient='records')
#criando arquivo excel
file=xlsx.load_workbook('planilha.xlsx')
allSheetNames=file.sheetnames
currentSheet=file['Planilha1']
for row in range (1, currentSheet.max_column+1, 1):
    for column in "ABC":
        cellName="{}".format(df['Resposta 1'])
#fim
