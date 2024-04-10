import openpyxl as xlsx
file=xlsx.load_workbook('planilha.xlsx')
allSheetNames=file.sheetnames
#print("Nomes das planilhas: {}".format(allSheetNames))
currentSheet=file['Planilha1']
for row in range (1, currentSheet.max_column+1, 1):
    for column in "ABC":
        cellName="{}{}".format(column, row)
        print("Posição:{} | Valor:{}".format(cellName,currentSheet[cellName].value))